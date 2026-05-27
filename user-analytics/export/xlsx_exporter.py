"""
export/xlsx_exporter.py
-----------------------
Geração de relatório em formato Excel (.xlsx) com formatação básica.
Utiliza openpyxl diretamente para controle fino sobre estilos.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config


# =============================================================================
# Constantes de estilo
# =============================================================================

_HEADER_FILL  = PatternFill("solid", fgColor="1A1A2E")
_HEADER_FONT  = Font(bold=True, color="00E5FF", name="Calibri", size=11)
_BORDER_SIDE  = Side(style="thin", color="DDDDDD")
_CELL_BORDER  = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE,  bottom=_BORDER_SIDE,
)
_ACTIVE_FILL   = PatternFill("solid", fgColor="E8F5E9")
_INACTIVE_FILL = PatternFill("solid", fgColor="FFF3E0")


# =============================================================================
# Função principal
# =============================================================================

def to_xlsx_bytes(rows: list[dict[str, Any]]) -> bytes:
    """Serializa uma lista de linhas do relatório para bytes .xlsx.

    Aplica formatação de cabeçalho, bordas, largura automática de colunas
    e destaca a linha com cor conforme o status do usuário.

    Parâmetros
    ----------
    rows:
        Lista de dicionários com as chaves definidas em config.REPORT_COLUMNS.

    Retorno
    -------
    Bytes do arquivo .xlsx prontos para envio via Flask.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Usuários"

    headers = config.REPORT_COLUMNS

    # ------------------------------------------------------------------
    # Cabeçalho
    # ------------------------------------------------------------------
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _CELL_BORDER

    ws.row_dimensions[1].height = 22

    # ------------------------------------------------------------------
    # Dados
    # ------------------------------------------------------------------
    for row_idx, row_data in enumerate(rows, start=2):
        is_active = str(row_data.get("status", "")).lower() == "ativo"
        row_fill  = _ACTIVE_FILL if is_active else _INACTIVE_FILL

        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key))
            cell.fill      = row_fill
            cell.border    = _CELL_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ------------------------------------------------------------------
    # Largura automática por coluna
    # ------------------------------------------------------------------
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    # ------------------------------------------------------------------
    # Serializa para bytes
    # ------------------------------------------------------------------
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
