"""
tests/test_exporter.py
----------------------
Testes unitários para export.csv_exporter e export.xlsx_exporter.

Verifica estrutura dos arquivos gerados, encoding, colunas e
comportamento com múltiplas linhas.
"""

from __future__ import annotations

import csv
import io

import pytest
from openpyxl import load_workbook

import config
from export.csv_exporter  import to_csv_bytes, build_report_row
from export.xlsx_exporter import to_xlsx_bytes


# =============================================================================
# Fixtures
# =============================================================================

@pytest.fixture
def single_row():
    return {
        "id_usuario":        1,
        "nome":              "Leanne Graham",
        "qtd_posts":         8,
        "media_chars":       150.5,
        "media_comentarios": 4.2,
        "status":            "Ativo",
    }


@pytest.fixture
def multi_rows():
    return [
        {"id_usuario": 1, "nome": "Alice",   "qtd_posts": 10,
         "media_chars": 200.0, "media_comentarios": 5.0, "status": "Ativo"},
        {"id_usuario": 2, "nome": "Bob",     "qtd_posts": 2,
         "media_chars": 80.0,  "media_comentarios": 2.0, "status": "Inativo"},
        {"id_usuario": 3, "nome": "Charlie", "qtd_posts": 6,
         "media_chars": 120.0, "media_comentarios": 3.5, "status": "Ativo"},
    ]


# =============================================================================
# Testes CSV
# =============================================================================

class TestCsvExporter:

    def test_returns_bytes(self, single_row):
        result = to_csv_bytes([single_row])
        assert isinstance(result, bytes)

    def test_utf8_bom_present(self, single_row):
        result = to_csv_bytes([single_row])
        assert result.startswith(b'\xef\xbb\xbf'), "BOM UTF-8 ausente"

    def test_header_matches_config_columns(self, single_row):
        csv_text = to_csv_bytes([single_row]).decode("utf-8-sig")
        reader   = csv.DictReader(io.StringIO(csv_text))
        assert reader.fieldnames == config.REPORT_COLUMNS

    def test_single_row_values(self, single_row):
        csv_text = to_csv_bytes([single_row]).decode("utf-8-sig")
        reader   = list(csv.DictReader(io.StringIO(csv_text)))
        assert len(reader) == 1
        assert reader[0]["nome"]       == "Leanne Graham"
        assert reader[0]["status"]     == "Ativo"
        assert reader[0]["qtd_posts"]  == "8"

    def test_multiple_rows(self, multi_rows):
        csv_text = to_csv_bytes(multi_rows).decode("utf-8-sig")
        reader   = list(csv.DictReader(io.StringIO(csv_text)))
        assert len(reader) == 3
        assert reader[1]["nome"] == "Bob"

    def test_empty_list_produces_only_header(self):
        csv_text = to_csv_bytes([]).decode("utf-8-sig")
        lines    = [l for l in csv_text.strip().splitlines() if l]
        assert len(lines) == 1  # só cabeçalho

    def test_comma_in_name_is_quoted(self):
        row = {
            "id_usuario": 1, "nome": "Smith, John",
            "qtd_posts": 5, "media_chars": 100.0,
            "media_comentarios": 3.0, "status": "Ativo",
        }
        csv_text = to_csv_bytes([row]).decode("utf-8-sig")
        assert '"Smith, John"' in csv_text


# =============================================================================
# Testes Excel
# =============================================================================

class TestXlsxExporter:

    def test_returns_bytes(self, single_row):
        result = to_xlsx_bytes([single_row])
        assert isinstance(result, bytes)

    def _load_ws(self, rows):
        xlsx_bytes = to_xlsx_bytes(rows)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        return wb.active

    def test_header_row_matches_config(self, single_row):
        ws      = self._load_ws([single_row])
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(config.REPORT_COLUMNS) + 1)]
        assert headers == config.REPORT_COLUMNS

    def test_single_row_values(self, single_row):
        ws = self._load_ws([single_row])
        assert ws.cell(row=2, column=2).value == "Leanne Graham"
        assert ws.cell(row=2, column=6).value == "Ativo"

    def test_multiple_rows_count(self, multi_rows):
        ws = self._load_ws(multi_rows)
        # header + 3 linhas de dados
        assert ws.max_row == 4

    def test_sheet_title(self, single_row):
        xlsx_bytes = to_xlsx_bytes([single_row])
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.active.title == "Relatório de Usuários"

    def test_empty_list_produces_only_header(self):
        ws = self._load_ws([])
        assert ws.max_row == 1


# =============================================================================
# Testes build_report_row
# =============================================================================

class TestBuildReportRow:

    def test_keys_match_config_columns(self):
        from analytics.metrics import compute_metrics
        posts = [{"id": 1, "body": "hello world", "_comment_count": 2}]
        m     = compute_metrics(posts, min_chars=0, min_posts=1)
        user  = {"id": 1, "name": "Test User"}
        row   = build_report_row(user, m)
        assert set(row.keys()) == set(config.REPORT_COLUMNS)

    def test_status_propagated_correctly(self):
        from analytics.metrics import compute_metrics
        posts = [{"id": 1, "body": "x" * 20, "_comment_count": 1}]
        m     = compute_metrics(posts, min_posts=1)
        row   = build_report_row({"id": 1, "name": "A"}, m)
        assert row["status"] == "Ativo"
